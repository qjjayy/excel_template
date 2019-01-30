#!/usr/bin/env python
# -*- coding:utf-8 -*-
from openpyxl.styles import Border
from copy import copy


class WriterTemplate(object):
    def __init__(self, sheet):
        self.sheet = sheet
        self.merged_cells = self.sheet.merged_cells
        self.row_size = len(list(sheet.rows))
        self.column_size = len(list(sheet.columns))
        self.row_index = 1
        self.template = self.get_template()

    def get_template(self):
        rows_config = list()
        top_config = rows_config
        for i in range(1, self.row_size + 1):
            # 诸列判断当前行是否为子模版的开始或结束
            continue_tag = False
            for j in range(1, self.column_size + 1):
                cell = self.sheet.cell(row=i, column=j)
                # 是否子模板开始
                if (isinstance(cell.value, (str, unicode)) and
                        cell.value.startswith('{{for ') and cell.value.endswith('}}')):
                    row_config = {
                        'list_field': cell.value[6:-2].strip(),
                        'sub_config': list()
                    }  # sub为list是为了兼容

                    top_config.append(row_config)
                    top_config = row_config['sub_config']
                    continue_tag = True
                    break
                # 是否子模板结束
                elif isinstance(cell.value, (str, unicode)) and cell.value == '{{end}}':
                    top_config = rows_config
                    continue_tag = True
                    break

            if continue_tag:
                continue

            row_config = {
                'columns_config': list(),
                'row_height': self.sheet.row_dimensions[i].height,
                'merged_range':
                    [{'min_col': cell_range.min_col, 'max_col': cell_range.max_col,
                      'min_row': cell_range.min_row, 'max_row': cell_range.max_row}
                     for cell_range in self.merged_cells if cell_range.min_row == i]
            }  # 列表项不支持跨行合并

            columns_config = row_config['columns_config']
            for j in range(1, self.column_size + 1):
                cell = self.sheet.cell(row=i, column=j)
                cell_config = self.process_cell(j, cell)
                if cell_config:
                    columns_config.append(cell_config)
            top_config.append(row_config)
        return rows_config

    def process_cell(self, column_index, cell):
        field = ''
        value = ''
        if (isinstance(cell.value, (str, unicode)) and
                cell.value.startswith('{{') and cell.value.endswith('}}')):
            field = cell.value[2:-2].strip()
            if field == '':
                field = 'null'  # 空值，手写单元格
        elif cell.value:
            value = cell.value
        else:
            return None  # 每个单元格必须有value或着{{}}，否则不会处理该单元格

        return {
            'value': value, 'field': field, 'col': column_index,
            'has_style': cell.has_style, 'font': cell.font,
            'border': cell.border, 'fill': cell.fill,
            'number_format': cell.number_format, 'protection': cell.protection,
            'alignment': cell.alignment, 'style': cell.style
        }

    def write(self, sheet, data):
        self.row_index = 1
        # 按照模版填写数据
        for row_config in self.template:
            if row_config.get('list_field'):
                loop_name = row_config['list_field']
                try:
                    loop_value = data[loop_name]
                except Exception:
                    raise Exception('Loop data({key}) lost!'.format(key=loop_name))
                self.write_sub(sheet, row_config['sub_config'], loop_value)
            else:
                self.write_row(sheet, row_config, data)
        # 填写剩余空格
        for i in range(self.row_index, self.row_size + 1):
            for j in range(1, self.column_size + 1):
                cell = sheet.cell(row=i, column=j)
                cell.value = None

    def write_sub(self, sheet, sub, loop_value):
        for value in loop_value:
            self.write_row(sheet, sub[0], value)

    def write_row(self, sheet, row_config, data):
        for cell_config in row_config['columns_config']:
            cell = sheet.cell(row=self.row_index, column=cell_config['col'])
            self.write_cell(cell, cell_config, data)
        # 设置行高和合并单元格
        sheet.row_dimensions[self.row_index].height = row_config['row_height']
        for cell_range in row_config['merged_range']:
            self.style_range(sheet, cell_range)
        self.row_index += 1

    def write_cell(self, cell, cell_config, data):
        if cell_config['value']:
            cell.value = cell_config['value']
        else:
            field_name = cell_config['field']
            try:
                if field_name == 'null':
                    value = ''
                else:
                    value = data[field_name]
            except Exception:
                raise Exception('data({key}) lost!'.format(key=field_name))
            cell.value = value

        if cell_config['has_style']:
            cell.font = copy(cell_config['font'])
            cell.border = copy(cell_config['border'])
            cell.fill = copy(cell_config['fill'])
            cell.number_format = cell_config['number_format']
            cell.protection = copy(cell_config['protection'])
            cell.alignment = copy(cell_config['alignment'])

    def style_range(self, sheet, cell_range):
        min_row = cell_range['min_row']
        max_row = cell_range['max_row']
        min_col = cell_range['min_col']
        max_col = cell_range['max_col']

        first_cell = sheet.cell(row=self.row_index, column=min_col)
        border = copy(first_cell.border)
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        if min_row != max_row:  # 仅支持表头
            sheet.merge_cells(
                start_row=min_row, start_column=min_col,
                end_row=max_row, end_column=max_col)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.border = cell.border + top
                    cell.border = cell.border + bottom
                    cell.border = cell.border + left
                    cell.border = cell.border + right
        else:
            sheet.merge_cells(
                start_row=self.row_index, start_column=cell_range['min_col'],
                end_row=self.row_index, end_column=cell_range['max_col'])

            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=self.row_index, column=col)
                cell.border = cell.border + top
                cell.border = cell.border + bottom

            left_cell = sheet.cell(row=self.row_index, column=min_col)
            left_cell.border = left_cell.border + left
            right_cell = sheet.cell(row=self.row_index, column=max_col)
            right_cell.border = right_cell.border + right
