#!/usr/bin/env python
# -*- coding:utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter
from WriterTemplate import WriterTemplate
try:
    import cStringIO as StringIO
except ImportError:
    import StringIO


class Writer(object):

    def __init__(self, template_file, sheet_name, out_file_path=None):
        self.template_file = template_file
        self.sheet_name = sheet_name

        if out_file_path:
            self.output = out_file_path
        else:
            self.output = StringIO.StringIO()
        
        self.template_workbook = load_workbook(self.template_file)
        self.out_workbook = Workbook()
        self.template_sheet = self.template_workbook[self.sheet_name]
        self.out_sheet = self.out_workbook.active
        self.writer_template = WriterTemplate(self.template_sheet)
        
    def __init_sheet(self, out_sheet):
        # 设置Sheet的总体样式
        out_sheet.page_breaks = self.template_sheet.page_breaks

        page_setup = self.template_sheet.page_setup
        page_setup.worksheet = out_sheet
        out_sheet.page_setup = self.template_sheet.page_setup

        out_sheet.print_options = self.template_sheet.print_options
        out_sheet.page_margins = self.template_sheet.page_margins
        out_sheet.protection = self.template_sheet.protection
        out_sheet.sheet_properties = self.template_sheet.sheet_properties
        out_sheet.sheet_format = self.template_sheet.sheet_format

        for i in range(1, len(list(self.template_sheet.columns)) + 1):
            column_name = get_column_letter(i)
            if self.template_sheet.column_dimensions[column_name].width:
                out_sheet.column_dimensions[column_name].width = \
                    self.template_sheet.column_dimensions[column_name].width
            else:
                out_sheet.column_dimensions[column_name].width = 8.38

    def set_data(self, data, multi_sheet=False):
        if multi_sheet:
            self.out_workbook.remove(self.out_workbook['Sheet'])
            for index, data_item in enumerate(data):
                sheet_name = 'Sheet' + str(index + 1)
                out_sheet = self.out_workbook.create_sheet(sheet_name)
                self.__init_sheet(out_sheet)
                self.writer_template.write(out_sheet, data_item)
        else:
            self.__init_sheet(self.out_sheet)
            self.writer_template.write(self.out_sheet, data)
        self.out_workbook.save(self.output)

        # 调试的时候使用
        # print ''
        # test_workbook = load_workbook(self.output)
        # for row in test_workbook.active.rows:
        #     for cell in row:
        #         print(cell.value),
        #     print ''

    def get_excel(self):
        self.output.seek(0)
        return self.output
